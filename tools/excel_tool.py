"""Excel tool — generate .xlsx spreadsheets."""
import os, uuid

NAME        = "excel_tool"
DESCRIPTION = "Generate an Excel (.xlsx) spreadsheet. Provide data as CSV-style rows or describe the table."
CATEGORY    = "builtin"
ICON        = "📊"
INPUTS = [
    {"name": "content",   "label": "Table Data (CSV format or description)", "type": "textarea",
     "placeholder": "Name,Age,Score\nAlice,25,92\nBob,30,87", "required": True},
    {"name": "sheet_name","label": "Sheet Name",     "type": "text",   "placeholder": "Sheet1"},
    {"name": "filename",  "label": "Output Filename","type": "text",   "placeholder": "data.xlsx"},
    {"name": "has_header","label": "First row is header?", "type": "select",
     "options": [{"value": "yes","label":"Yes"},{"value":"no","label":"No"}]},
]

DOCS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "generated_docs")


def run(content: str, sheet_name: str = "Sheet1", filename: str = "",
        has_header: str = "yes") -> dict:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import csv, io

        os.makedirs(DOCS_DIR, exist_ok=True)
        if not filename:
            filename = f"data_{uuid.uuid4().hex[:8]}.xlsx"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        path = os.path.join(DOCS_DIR, filename)

        # Parse CSV content
        reader = csv.reader(io.StringIO(content.strip()))
        rows   = [row for row in reader if any(c.strip() for c in row)]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name or "Sheet1"

        header_fill  = PatternFill("solid", fgColor="1a73e8")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        alt_fill     = PatternFill("solid", fgColor="EEF4FF")
        border_side  = Side(style="thin", color="CCCCCC")
        cell_border  = Border(left=border_side, right=border_side,
                              top=border_side,  bottom=border_side)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val.strip())
                cell.border   = cell_border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if r_idx == 1 and has_header == "yes":
                    cell.fill = header_fill
                    cell.font = header_font
                elif r_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Freeze header row
        if has_header == "yes" and rows:
            ws.freeze_panes = "A2"

        wb.save(path)
        return {"message": f"✅ Excel spreadsheet generated! ({len(rows)} rows)\n[📥 Download {filename}](/docs/{filename})",
                "file": f"/docs/{filename}"}

    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}
    except Exception as e:
        return {"error": f"Excel generation failed: {e}"}
